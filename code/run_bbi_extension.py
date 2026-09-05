#!/usr/bin/env python3
"""Reconstruct biological sample level endothelial results used in the BBI manuscript.

The workflow calculates animal and donor summaries, acute EAE meta analysis,
leave one dataset out sensitivity estimates, cross species gene effects, lesion
stage summaries and the three group VEGF A transcriptional comparison.
"""

from __future__ import annotations

import csv
import gzip
import io
import itertools
import json
import math
import re
import tarfile
import tempfile
import zipfile
from pathlib import Path

import h5py
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy import sparse
from scipy.io import mmread
from scipy.stats import binomtest, mannwhitneyu, norm, spearmanr, t


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "public_data"
SOURCE = ROOT / "supplementary_tables"
OUT = ROOT / "results"
SEED = 20260901
RNG = np.random.default_rng(SEED)


# The three state modules were fixed before the extension analyses.  Human and
# mouse orthologues are represented by upper-case symbols.  Species-specific
# MHC genes are both retained; missing genes are ignored within each dataset.
STATE_MODULES = {
    "Endothelial_immune_activation": [
        "STAT1", "STAT3", "IRF1", "CXCL9", "CXCL10", "CCL2", "B2M",
        "TAP1", "HLA-DRA", "HLA-DRB1", "H2-AA", "H2-AB1", "ICAM1",
        "VCAM1", "SELE", "SELP", "IFIT1", "IFIT2", "IFIT3", "ISG15",
    ],
    "BBB_specialization": [
        "CLDN5", "OCLN", "TJP1", "LSR", "MFSD2A", "SLC2A1", "ABCG2",
        "ABCB1", "LEF1", "AXIN2", "APCDD1", "ADGRA2", "RECK", "TCF7L2",
    ],
    "Structural_ECM_remodeling": [
        "CAV1", "CAV2", "EHD2", "PLVAP", "MMP2", "MMP9", "TIMP1",
        "TIMP2", "COL4A1", "COL4A2", "FN1", "LAMA4", "LAMB1", "SPARC",
        "VIM",
    ],
}

LOCKED_MODULES = {
    "Tight_junction": ["CLDN5", "OCLN", "TJP1", "LSR"],
    "BBB_transport_identity": ["MFSD2A", "SLC2A1", "ABCG2", "ABCB1"],
    "Caveolae_structural": ["CAV1", "CAV2", "EHD2"],
    "PLVAP_permeability": ["PLVAP", "ESM1", "APLN", "ANGPT2"],
    "Adhesion_trafficking": ["ICAM1", "VCAM1", "SELE", "SELP", "CCL2", "CXCL10"],
    "IFN_antigen_presentation": [
        "STAT1", "IRF1", "CXCL9", "CXCL10", "B2M", "TAP1", "HLA-DRA",
        "HLA-DRB1", "H2-AA", "H2-AB1", "IFIT1", "IFIT2", "IFIT3", "ISG15",
    ],
    "Wnt_BBB": ["LEF1", "AXIN2", "APCDD1", "ADGRA2", "RECK", "TCF7L2"],
    "TGF_response": ["TGFBR1", "TGFBR2", "SMAD2", "SMAD3", "SMAD4", "SMAD7", "SERPINE1", "CTGF"],
    "VEGF_response": ["KDR", "FLT1", "PGF", "ANGPT2", "ESM1", "APLN"],
    "ROS_Src": ["SRC", "RAC1", "NOX4", "CYBB", "NCF1", "NCF2", "HIF1A"],
    "ECM_protease": ["MMP2", "MMP9", "TIMP1", "TIMP2"],
}

ALL_MODULES = {**STATE_MODULES, **LOCKED_MODULES}

EC_CORE = [
    "PECAM1", "VWF", "EMCN", "CDH5", "RAMP2", "RGCC", "CA4", "KDR",
    "FLT1", "TEK", "ENG", "ESAM",
]
MURAL = ["RGS5", "PDGFRB", "CSPG4", "ABCC9", "KCNJ8", "DES", "ACTA2"]
GLIAL = [
    "MBP", "PLP1", "MOG", "MOBP", "GFAP", "AQP4", "ALDH1L1", "SLC1A2",
    "P2RY12", "C1QA", "C1QB", "AIF1", "TYROBP",
]
IMMUNE = ["PTPRC", "CD3D", "CD3E", "MS4A1", "CD79A", "NKG7", "LYZ"]


def ensure_out() -> None:
    OUT.mkdir(parents=True, exist_ok=True)


def upper_gene(x: object) -> str:
    return str(x).strip().upper().replace("_", "-")


def aggregate_duplicate_genes(genes: list[str], values: np.ndarray) -> tuple[list[str], np.ndarray]:
    """Sum duplicate gene rows for a vector or gene-by-sample matrix."""
    genes_u = np.array([upper_gene(g) for g in genes], dtype=object)
    keep = genes_u != ""
    genes_u = genes_u[keep]
    arr = np.asarray(values)[keep]
    unique, inv = np.unique(genes_u, return_inverse=True)
    if arr.ndim == 1:
        out = np.zeros(len(unique), dtype=float)
        np.add.at(out, inv, arr.astype(float))
    else:
        out = np.zeros((len(unique), arr.shape[1]), dtype=float)
        for i in range(arr.shape[0]):
            out[inv[i]] += arr[i]
    return unique.tolist(), out


def log_cpm(counts: np.ndarray) -> np.ndarray:
    counts = np.asarray(counts, dtype=float)
    if counts.ndim == 1:
        return np.log1p(counts / max(float(counts.sum()), 1.0) * 1e6)
    lib = counts.sum(axis=0)
    return np.log1p(counts / np.maximum(lib, 1.0)[None, :] * 1e6)


def zscore_rows(a: np.ndarray) -> np.ndarray:
    a = np.asarray(a, dtype=float)
    mu = np.nanmean(a, axis=1, keepdims=True)
    sd = np.nanstd(a, axis=1, ddof=1, keepdims=True)
    sd[(~np.isfinite(sd)) | (sd == 0)] = 1.0
    return (a - mu) / sd


def module_scores(
    genes: list[str], expression: np.ndarray, sample_ids: list[str], modules: dict[str, list[str]] = ALL_MODULES
) -> tuple[pd.DataFrame, pd.DataFrame]:
    genes_u = np.array([upper_gene(g) for g in genes])
    lookup = {g: i for i, g in enumerate(genes_u)}
    z = zscore_rows(expression)
    scores: dict[str, np.ndarray] = {}
    coverage = []
    for name, members in modules.items():
        present = [g for g in members if g in lookup]
        idx = [lookup[g] for g in present]
        scores[name] = np.nanmean(z[idx], axis=0) if idx else np.full(expression.shape[1], np.nan)
        coverage.append(
            {
                "module": name,
                "genes_expected": len(members),
                "genes_found": len(present),
                "genes_used": ";".join(present),
            }
        )
    out = pd.DataFrame(scores)
    out.insert(0, "sample_id", sample_ids)
    return out, pd.DataFrame(coverage)


def bh_fdr(pvalues: list[float] | np.ndarray) -> np.ndarray:
    p = np.asarray(pvalues, dtype=float)
    ans = np.full_like(p, np.nan)
    ok = np.isfinite(p)
    if not ok.any():
        return ans
    x = p[ok]
    order = np.argsort(x)
    ranked = x[order] * len(x) / np.arange(1, len(x) + 1)
    ranked = np.minimum.accumulate(ranked[::-1])[::-1]
    inv = np.empty_like(order)
    inv[order] = np.arange(len(order))
    ans[ok] = np.minimum(ranked[inv], 1.0)
    return ans


def exact_mean_permutation(x: np.ndarray, y: np.ndarray, alternative: str = "two-sided") -> tuple[float, float]:
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    observed = float(np.mean(x) - np.mean(y))
    combined = np.r_[x, y]
    n_x = len(x)
    total = 0
    extreme = 0
    for idx in itertools.combinations(range(len(combined)), n_x):
        mask = np.zeros(len(combined), dtype=bool)
        mask[list(idx)] = True
        stat = float(np.mean(combined[mask]) - np.mean(combined[~mask]))
        total += 1
        if alternative == "greater":
            extreme += stat >= observed - 1e-12
        elif alternative == "less":
            extreme += stat <= observed + 1e-12
        else:
            extreme += abs(stat) >= abs(observed) - 1e-12
    return observed, extreme / total


def cliffs_delta(x: np.ndarray, y: np.ndarray) -> float:
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    return float((np.sum(x[:, None] > y[None, :]) - np.sum(x[:, None] < y[None, :])) / (len(x) * len(y)))


def hedges_g(x: np.ndarray, y: np.ndarray) -> tuple[float, float, float, float]:
    """Hedges g for x minus y, its sampling variance, and Wald CI."""
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    nx, ny = len(x), len(y)
    df = nx + ny - 2
    if df <= 0:
        return np.nan, np.nan, np.nan, np.nan
    pooled_var = ((nx - 1) * np.var(x, ddof=1) + (ny - 1) * np.var(y, ddof=1)) / df
    if pooled_var <= 0 or not np.isfinite(pooled_var):
        return np.nan, np.nan, np.nan, np.nan
    d = (float(np.mean(x)) - float(np.mean(y))) / math.sqrt(pooled_var)
    correction = 1.0 - 3.0 / (4.0 * df - 1.0)
    g = correction * d
    variance = (nx + ny) / (nx * ny) + (g * g) / (2.0 * df)
    se = math.sqrt(variance)
    return g, variance, g - 1.96 * se, g + 1.96 * se


def contrast_rows(
    scores: pd.DataFrame,
    group_col: str,
    case: str,
    control: str,
    dataset: str,
    contrast_label: str,
    modules: dict[str, list[str]] = STATE_MODULES,
) -> list[dict[str, object]]:
    rows = []
    for module in modules:
        x = scores.loc[scores[group_col] == case, module].dropna().to_numpy(float)
        y = scores.loc[scores[group_col] == control, module].dropna().to_numpy(float)
        if len(x) < 2 or len(y) < 2:
            continue
        delta, p = exact_mean_permutation(x, y)
        g, var_g, lo, hi = hedges_g(x, y)
        rows.append(
            {
                "dataset": dataset,
                "contrast": contrast_label,
                "module": module,
                "case": case,
                "control": control,
                "n_case": len(x),
                "n_control": len(y),
                "mean_difference": delta,
                "p_exact": p,
                "cliffs_delta": cliffs_delta(x, y),
                "hedges_g": g,
                "var_g": var_g,
                "ci_low": lo,
                "ci_high": hi,
            }
        )
    q = bh_fdr([r["p_exact"] for r in rows])
    for row, fdr in zip(rows, q):
        row["FDR_BH"] = fdr
    return rows


def read_10x_tar_matrix(path: Path) -> tuple[list[str], sparse.csc_matrix, int]:
    with tarfile.open(path, "r:gz") as tar:
        names = tar.getnames()
        matrix_name = next(n for n in names if n.endswith("matrix.mtx.gz") and "/._" not in n)
        feature_name = next(n for n in names if n.endswith("features.tsv.gz") and "/._" not in n)
        barcode_name = next(n for n in names if n.endswith("barcodes.tsv.gz") and "/._" not in n)
        matrix_raw = tar.extractfile(matrix_name)
        feature_raw = tar.extractfile(feature_name)
        barcode_raw = tar.extractfile(barcode_name)
        assert matrix_raw is not None and feature_raw is not None and barcode_raw is not None
        with gzip.GzipFile(fileobj=matrix_raw, mode="rb") as fh:
            matrix = mmread(fh).tocsc()
        with gzip.GzipFile(fileobj=feature_raw, mode="rb") as fh:
            features = pd.read_csv(fh, sep="\t", header=None)
        with gzip.GzipFile(fileobj=barcode_raw, mode="rb") as fh:
            n_cells = len(pd.read_csv(fh, sep="\t", header=None))
    gene_col = 1 if features.shape[1] > 1 else 0
    return features.iloc[:, gene_col].astype(str).tolist(), matrix, n_cells


def gene_payload(
    dataset: str,
    genes: list[str],
    expression: np.ndarray,
    sample_ids: list[str],
    groups: list[str],
    case: str,
    control: str,
) -> dict[str, object]:
    """Keep animal-level gene expression in memory for cross-species analysis."""
    return {
        "dataset": dataset,
        "genes": [upper_gene(g) for g in genes],
        "expression": np.asarray(expression, dtype=float),
        "sample_ids": list(sample_ids),
        "groups": list(groups),
        "case": case,
        "control": control,
    }


def analyze_gse210776() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, object]]:
    folder = DATA / "GSE210776"
    group_map = {
        "DS001": "CFA", "DS002": "Acute", "DS003": "Chronic", "DS004": "CFA",
        "DS005": "Acute", "DS006": "Chronic", "DS007": "Healthy", "DS008": "Healthy",
        "DS009": "Chronic", "DS010": "Chronic", "DS011": "Acute", "DS012": "Acute",
        "DS013": "Healthy", "DS014": "CFA", "BEVAC001": "Bevacizumab",
        "BEVAC002": "Bevacizumab", "BEVAC003": "Bevacizumab",
    }
    vectors: list[np.ndarray] = []
    sample_ids: list[str] = []
    groups: list[str] = []
    genes_ref: list[str] | None = None
    qc = []
    for path in sorted(folder.glob("*.tar.gz")):
        match = re.search(r"_(DS\d+|BEVAC\d+)_processed", path.name)
        if not match:
            continue
        sample = match.group(1)
        genes, matrix, n_barcodes = read_10x_tar_matrix(path)
        # The GEO archives contain Cell Ranger raw droplet matrices rather than
        # filtered cell matrices.  First remove low-count droplets, then retain
        # endothelial cells using markers that are independent of the outcome
        # modules.  Summing every barcode would otherwise fold ambient RNA into
        # the animal-level profiles.
        totals = np.asarray(matrix.sum(axis=0)).ravel()
        detected = np.asarray(matrix.getnnz(axis=0)).ravel()
        good_qc = (totals >= 300) & (detected >= 150)
        genes_u = np.array([upper_gene(g) for g in genes])
        lookup = {g: i for i, g in enumerate(genes_u)}
        target = sorted(set(EC_CORE + MURAL + GLIAL + IMMUNE))
        present = [g for g in target if g in lookup]
        target_expr = np.log1p(
            np.asarray(matrix[[lookup[g] for g in present], :][:, good_qc].toarray(), dtype=float)
            / np.maximum(totals[good_qc], 1.0)[None, :]
            * 1e4
        )
        target_df = pd.DataFrame(target_expr, index=present)

        def marker_score(members: list[str]) -> np.ndarray:
            observed = [g for g in members if g in target_df.index]
            return target_df.loc[observed].mean(axis=0).to_numpy() if observed else np.zeros(good_qc.sum())

        ec_present = [g for g in EC_CORE if g in lookup]
        ec_hits = np.asarray((matrix[[lookup[g] for g in ec_present], :][:, good_qc] > 0).sum(axis=0)).ravel()
        ec_score = marker_score(EC_CORE)
        contaminant = np.maximum.reduce([marker_score(MURAL), marker_score(GLIAL), marker_score(IMMUNE)])
        selected_within_qc = (ec_hits >= 2) & (ec_score > contaminant)
        selected = np.where(good_qc)[0][selected_within_qc]
        counts = np.asarray(matrix[:, selected].sum(axis=1)).ravel()
        genes, counts = aggregate_duplicate_genes(genes, counts)
        if genes_ref is None:
            genes_ref = genes
        elif genes != genes_ref:
            lookup = dict(zip(genes, counts))
            counts = np.array([lookup.get(g, 0.0) for g in genes_ref])
        vectors.append(counts)
        sample_ids.append(sample)
        groups.append(group_map[sample])
        qc.append(
            {
                "dataset": "GSE210776",
                "sample_id": sample,
                "group": group_map[sample],
                "raw_barcodes": n_barcodes,
                "barcodes_after_qc": int(good_qc.sum()),
                "endothelial_cells": int(len(selected)),
                "selection_rule": "UMI>=300;genes>=150;EC_hits>=2;EC_score>contaminant",
                "total_UMI": float(counts.sum()),
            }
        )
    assert genes_ref is not None
    expr = log_cpm(np.column_stack(vectors))
    scores, coverage = module_scores(genes_ref, expr, sample_ids)
    scores.insert(1, "group", groups)
    scores.insert(1, "dataset", "GSE210776")
    coverage.insert(0, "dataset", "GSE210776")
    payload = gene_payload("GSE210776", genes_ref, expr, sample_ids, groups, "Acute", "CFA")
    return scores, coverage, pd.DataFrame(qc), payload


def analyze_gse199460() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, object]]:
    folder = DATA / "GSE199460"
    meta = pd.read_csv(folder / "GSE199460_cell_annotation.meta_data.cd31_selection.csv.gz", index_col=0)
    meta["matrix_barcode"] = meta.index.str.replace("-1$", ".1", regex=True)
    expr_path = folder / "GSE199460_normalized_expr.cd31_selection.sctransform.csv.gz"
    header = pd.read_csv(expr_path, nrows=0).columns.astype(str)
    matrix_columns = header[1:] if header[0].startswith("Unnamed") else header
    # The expression file has a blank first header cell, which pandas uses as the row index.
    matrix_columns = pd.read_csv(expr_path, nrows=0, index_col=0).columns.astype(str)
    column_meta = meta.set_index("matrix_barcode").reindex(matrix_columns)
    valid = column_meta["orig.ident"].notna().to_numpy()
    sample_order = sorted(column_meta.loc[valid, "orig.ident"].unique())
    all_groups = {s: np.where(column_meta["orig.ident"].eq(s).to_numpy())[0] for s in sample_order}
    venous_mask = column_meta["cell_types"].isin(["venous", "cap venous"]).fillna(False).to_numpy()
    venous_groups = {s: idx[venous_mask[idx]] for s, idx in all_groups.items()}
    all_blocks: list[np.ndarray] = []
    venous_blocks: list[np.ndarray] = []
    genes: list[str] = []
    for chunk in pd.read_csv(expr_path, index_col=0, chunksize=250):
        arr = chunk.to_numpy(float)
        all_blocks.append(np.column_stack([np.nanmean(arr[:, idx], axis=1) for idx in all_groups.values()]))
        venous_blocks.append(np.column_stack([np.nanmean(arr[:, idx], axis=1) for idx in venous_groups.values()]))
        genes.extend(chunk.index.astype(str).tolist())
    condition = ["Acute" if s.startswith("EAE") else "Control" for s in sample_order]
    outputs = []
    coverage_parts = []
    all_expression = np.vstack(all_blocks)
    for subset, blocks in [("All_endothelial", all_blocks), ("Venous_like", venous_blocks)]:
        expression = np.vstack(blocks)
        scores, coverage = module_scores(genes, expression, sample_order)
        scores.insert(1, "group", condition)
        scores.insert(1, "subset", subset)
        scores.insert(1, "dataset", "GSE199460")
        coverage.insert(0, "subset", subset)
        coverage.insert(0, "dataset", "GSE199460")
        outputs.append(scores)
        coverage_parts.append(coverage)
    qc = meta.groupby(["orig.ident", "condition", "cell_types"]).size().rename("cells").reset_index()
    qc.insert(0, "dataset", "GSE199460")
    payload = gene_payload("GSE199460", genes, all_expression, sample_order, condition, "Acute", "Control")
    return pd.concat(outputs, ignore_index=True), pd.concat(coverage_parts, ignore_index=True), qc, payload


def analyze_gse95401() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, object]]:
    path = DATA / "GSE95401" / "GSE95401_MunjiSoungDaneman_RNAseq_BBB_Health_Disease.xlsx"
    ws = load_workbook(path, read_only=True, data_only=True)["Replicates and Averages"]
    rows = ws.iter_rows(values_only=True)
    header = [str(x).strip() if x is not None else "" for x in next(rows)]
    wanted = [i for i, x in enumerate(header) if re.fullmatch(r"EAE (Control|Acute|Subacute|Chronic)\d\s*", x)]
    sample_ids = [header[i].replace(" ", "_") for i in wanted]
    groups = [re.match(r"EAE (Control|Acute|Subacute|Chronic)", header[i]).group(1) for i in wanted]
    genes = []
    values = []
    for row in rows:
        gene = row[0]
        if gene is None:
            continue
        vals = [row[i] for i in wanted]
        if not any(v is not None for v in vals):
            continue
        genes.append(str(gene))
        values.append([float(v or 0.0) for v in vals])
    genes, counts = aggregate_duplicate_genes(genes, np.asarray(values, dtype=float))
    expr = log_cpm(counts)
    scores, coverage = module_scores(genes, expr, sample_ids)
    scores.insert(1, "group", groups)
    scores.insert(1, "dataset", "GSE95401")
    coverage.insert(0, "dataset", "GSE95401")
    qc = pd.DataFrame({"dataset": "GSE95401", "sample_id": sample_ids, "group": groups, "total_count": counts.sum(axis=0)})
    payload = gene_payload("GSE95401", genes, expr, sample_ids, groups, "Acute", "Control")
    return scores, coverage, qc, payload


def read_mtx_triplet(folder: Path, prefix: str) -> tuple[list[str], pd.Series, sparse.csc_matrix]:
    matrix_file = next(folder.glob(prefix + "*_matrix.mtx.gz"))
    feature_file = next(folder.glob(prefix + "*_features.tsv.gz"))
    barcode_file = next(folder.glob(prefix + "*_barcodes.tsv.gz"))
    with gzip.open(matrix_file, "rb") as fh:
        matrix = mmread(fh).tocsc()
    features = pd.read_csv(feature_file, sep="\t", header=None)
    barcodes = pd.read_csv(barcode_file, sep="\t", header=None).iloc[:, 0].astype(str)
    gene_col = 1 if features.shape[1] > 1 else 0
    return features.iloc[:, gene_col].astype(str).tolist(), barcodes, matrix


def analyze_gse279183() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, object]]:
    folder = DATA / "GSE279183_snRNA"
    sample_specs = [
        ("GSM8563681_CO37", "CO37", "CO37", "Control"),
        ("GSM8563682_CO40", "CO40", "CO40", "Control"),
        ("GSM8563683_CO41", "CO41", "CO41", "Control"),
        ("GSM8563684_CO45", "CO45", "CO45", "Control"),
        ("GSM8563685_CO74", "CO74", "CO74", "Control"),
        ("GSM8563686_CO85", "CO85", "CO85", "Control"),
        ("GSM8563687_MS197D", "MS197D", "MS197", "Chronic_active"),
        ("GSM8563688_MS229", "MS229", "MS229", "Chronic_active"),
        ("GSM8563689_MS377N", "MS377N", "MS377", "Chronic_active"),
        ("GSM8563690_MS377T", "MS377T", "MS377", "Chronic_active"),
        ("GSM8563691_MS377I", "MS377I", "MS377", "Chronic_active"),
        ("GSM8563692_MS411", "MS411", "MS411", "Chronic_active"),
        ("GSM8563693_MS497I", "MS497I", "MS497", "Chronic_inactive"),
        ("GSM8563694_MS497T", "MS497T", "MS497", "Chronic_inactive"),
        ("GSM8563695_MS549H", "MS549H", "MS549", "Chronic_inactive"),
        ("GSM8563696_MS549T", "MS549T", "MS549", "Chronic_inactive"),
    ]
    target = sorted(set(EC_CORE + MURAL + GLIAL + IMMUNE + sum(ALL_MODULES.values(), [])))
    tissue_vectors = []
    genes_ref = None
    tissue_meta = []
    qc = []
    for prefix, tissue, donor, group in sample_specs:
        genes, barcodes, matrix = read_mtx_triplet(folder, prefix)
        genes_u = np.array([upper_gene(g) for g in genes])
        lookup = {g: i for i, g in enumerate(genes_u)}
        totals = np.asarray(matrix.sum(axis=0)).ravel()
        detected = np.asarray(matrix.getnnz(axis=0)).ravel()
        idx = [lookup[g] for g in target if g in lookup]
        names = [genes_u[i] for i in idx]
        norm_target = np.log1p(np.asarray(matrix[idx].toarray(), dtype=float) / np.maximum(totals, 1.0)[None, :] * 1e4)
        ndf = pd.DataFrame(norm_target, index=names)

        def score(members: list[str]) -> np.ndarray:
            present = [g for g in members if g in ndf.index]
            return ndf.loc[present].mean(axis=0).to_numpy() if present else np.zeros(matrix.shape[1])

        core_present = [g for g in EC_CORE if g in lookup]
        ec_hits = np.asarray((matrix[[lookup[g] for g in core_present]] > 0).sum(axis=0)).ravel()
        ec_score = score(EC_CORE)
        contaminant = np.maximum.reduce([score(MURAL), score(GLIAL), score(IMMUNE)])
        good_qc = (totals >= 300) & (detected >= 150)
        selected = good_qc & (ec_hits >= 2) & (ec_score > contaminant)
        # A permissive fall-back is declared and recorded if a very small tissue
        # fragment contains fewer than 20 endothelial nuclei.
        rule = "EC_hits>=2_and_EC_score>contaminant"
        if selected.sum() < 20:
            selected = good_qc & (ec_hits >= 2)
            rule = "fallback_EC_hits>=2"
        counts = np.asarray(matrix[:, selected].sum(axis=1)).ravel()
        genes_agg, counts_agg = aggregate_duplicate_genes(genes, counts)
        if genes_ref is None:
            genes_ref = genes_agg
        elif genes_agg != genes_ref:
            d = dict(zip(genes_agg, counts_agg))
            counts_agg = np.array([d.get(g, 0.0) for g in genes_ref])
        tissue_vectors.append(counts_agg)
        tissue_meta.append({"tissue": tissue, "donor": donor, "group": group})
        qc.append(
            {
                "dataset": "GSE279183",
                "tissue": tissue,
                "donor": donor,
                "group": group,
                "nuclei_total": matrix.shape[1],
                "nuclei_qc": int(good_qc.sum()),
                "endothelial_nuclei": int(selected.sum()),
                "selection_rule": rule,
                "median_UMI_selected": float(np.median(totals[selected])) if selected.any() else np.nan,
            }
        )
    assert genes_ref is not None
    tissue_counts = np.column_stack(tissue_vectors)
    meta = pd.DataFrame(tissue_meta)
    donor_order = meta["donor"].drop_duplicates().tolist()
    donor_counts = []
    donor_group = []
    for donor in donor_order:
        cols = np.where(meta["donor"].eq(donor).to_numpy())[0]
        donor_counts.append(tissue_counts[:, cols].sum(axis=1))
        donor_group.append(meta.loc[cols[0], "group"])
    expr = log_cpm(np.column_stack(donor_counts))
    scores, coverage = module_scores(genes_ref, expr, donor_order)
    scores.insert(1, "group", donor_group)
    scores.insert(1, "dataset", "GSE279183")
    coverage.insert(0, "dataset", "GSE279183")
    payload = gene_payload(
        "GSE279183", genes_ref, expr, donor_order, donor_group, "Chronic_active", "Control"
    )
    return scores, coverage, pd.DataFrame(qc), payload


def read_10x_h5_from_tar(path: Path, sample: str) -> tuple[list[str], pd.Series, sparse.csc_matrix, pd.DataFrame]:
    with tarfile.open(path, "r:gz") as tar, tempfile.TemporaryDirectory() as tmp:
        h5_member = tar.getmember(f"{sample}/filtered_feature_bc_matrix.h5")
        h5_member.uid = h5_member.gid = 0
        tar.extract(h5_member, tmp, filter="data")
        with h5py.File(Path(tmp) / sample / "filtered_feature_bc_matrix.h5", "r") as h:
            group = h["matrix"]
            shape = tuple(int(x) for x in group["shape"][:])
            matrix = sparse.csc_matrix(
                (group["data"][:], group["indices"][:], group["indptr"][:]), shape=shape
            )
            genes = [x.decode() for x in group["features"]["name"][:]]
            barcodes = pd.Series([x.decode() for x in group["barcodes"][:]])
        pos_member = tar.getmember(f"{sample}/spatial/tissue_positions_list.csv")
        pos_raw = tar.extractfile(pos_member)
        assert pos_raw is not None
        positions = pd.read_csv(
            pos_raw,
            header=None,
            names=["barcode", "in_tissue", "array_row", "array_col", "pxl_row", "pxl_col"],
        )
    return genes, barcodes, matrix, positions


def analyze_gse208747() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    folder = DATA / "GSE208747"
    group_map = {"A": "Active", "C": "Control", "M": "Mixed_active_inactive", "N": "NAWM"}
    donor_map = {
        "A1": "Active_1", "A3": "Active_1", "A2": "Active_2", "A4": "Active_3",
        "C1": "Control_1", "C2": "Control_2",
        "M1": "Mixed_1", "M2": "Mixed_2", "M3": "Mixed_3", "M4": "Mixed_4", "M5": "Mixed_5", "M6": "Mixed_6",
        "N1": "NAWM_1", "N2": "NAWM_2", "N3": "NAWM_3",
    }
    target = sorted(set(EC_CORE + sum(ALL_MODULES.values(), [])))
    spot_frames = []
    sample_pseudobulk = []
    sample_ids = []
    qc = []
    coverage_rows = []
    for path in sorted(folder.glob("*.tar.gz")):
        sample = re.search(r"_([ACMN]\d+)\.tar\.gz$", path.name).group(1)
        genes, barcodes, matrix, positions = read_10x_h5_from_tar(path, sample)
        positions = positions.set_index("barcode").reindex(barcodes)
        in_tissue = positions["in_tissue"].fillna(0).astype(int).to_numpy() == 1
        totals = np.asarray(matrix.sum(axis=0)).ravel()
        detected = np.asarray(matrix.getnnz(axis=0)).ravel()
        good = in_tissue & (totals >= 500) & (detected >= 200)
        genes_u = np.array([upper_gene(g) for g in genes])
        lookup = {g: i for i, g in enumerate(genes_u)}
        idx = [lookup[g] for g in target if g in lookup]
        names = [genes_u[i] for i in idx]
        lognorm = np.log1p(
            np.asarray(matrix[idx, :][:, good].toarray(), dtype=float)
            / np.maximum(totals[good], 1.0)[None, :]
            * 1e4
        )
        target_df = pd.DataFrame(lognorm.T, columns=names)
        ec_present = [g for g in EC_CORE if g in target_df.columns]
        ec_score = target_df[ec_present].mean(axis=1)
        cutoff = float(ec_score.quantile(0.70))
        selected = ec_score >= cutoff
        frame = target_df.loc[selected].copy()
        frame.insert(0, "EC_score", ec_score.loc[selected].to_numpy())
        frame.insert(0, "sample_id", sample)
        spot_frames.append(frame)
        selected_good_idx = np.where(good)[0][selected.to_numpy()]
        counts = np.asarray(matrix[:, selected_good_idx].sum(axis=1)).ravel()
        genes_agg, counts_agg = aggregate_duplicate_genes(genes, counts)
        sample_pseudobulk.append((genes_agg, counts_agg))
        sample_ids.append(sample)
        qc.append(
            {
                "dataset": "GSE208747",
                "sample_id": sample,
                "donor": donor_map[sample],
                "group": group_map[sample[0]],
                "spots_under_tissue": int(in_tissue.sum()),
                "spots_after_qc": int(good.sum()),
                "EC_rich_spots_q30": int(selected.sum()),
                "EC_score_cutoff": cutoff,
            }
        )
    all_spots = pd.concat(spot_frames, ignore_index=True)
    # Standardize genes across all endothelial-rich spots, calculate modules,
    # then residualize each module against the pan-endothelial score.
    gene_columns = [c for c in all_spots.columns if c not in {"sample_id", "EC_score"}]
    z = zscore_rows(all_spots[gene_columns].to_numpy(float).T).T
    zdf = pd.DataFrame(z, columns=gene_columns)
    spot_scores = pd.DataFrame({"sample_id": all_spots["sample_id"], "EC_score": all_spots["EC_score"]})
    for module, members in ALL_MODULES.items():
        present = [g for g in members if g in zdf.columns]
        raw = zdf[present].mean(axis=1).to_numpy() if present else np.full(len(zdf), np.nan)
        x = np.column_stack([np.ones(len(raw)), spot_scores["EC_score"].to_numpy(float)])
        beta, *_ = np.linalg.lstsq(x, raw, rcond=None)
        spot_scores[module] = raw - x @ beta
        coverage_rows.append({"dataset": "GSE208747", "module": module, "genes_expected": len(members), "genes_found": len(present), "genes_used": ";".join(present)})
    sample_scores = spot_scores.groupby("sample_id")[list(ALL_MODULES)].median().reset_index()
    sample_scores["donor"] = sample_scores["sample_id"].map(donor_map)
    sample_scores["group"] = sample_scores["sample_id"].str[0].map(group_map)
    donor_scores = sample_scores.groupby(["donor", "group"])[list(ALL_MODULES)].mean().reset_index()
    donor_scores = donor_scores.rename(columns={"donor": "sample_id"})
    donor_scores.insert(1, "dataset", "GSE208747")
    return donor_scores, pd.DataFrame(coverage_rows).drop_duplicates(), pd.DataFrame(qc)


def random_effects_meta(effect_table: pd.DataFrame, label: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    pooled_rows = []
    loo_rows = []
    for module, block in effect_table.groupby("module"):
        block = block[np.isfinite(block["hedges_g"]) & np.isfinite(block["var_g"])].copy()
        if len(block) < 2:
            continue

        def pool(b: pd.DataFrame) -> dict[str, float]:
            g = b["hedges_g"].to_numpy(float)
            v = b["var_g"].to_numpy(float)
            w = 1.0 / v
            fixed = float(np.sum(w * g) / np.sum(w))
            q = float(np.sum(w * (g - fixed) ** 2))
            df = len(g) - 1
            c = float(np.sum(w) - np.sum(w * w) / np.sum(w))
            tau2 = max(0.0, (q - df) / c) if c > 0 else 0.0
            wr = 1.0 / (v + tau2)
            est = float(np.sum(wr * g) / np.sum(wr))
            se = math.sqrt(1.0 / float(np.sum(wr)))
            i2 = max(0.0, (q - df) / q) * 100.0 if q > 0 else 0.0
            # Modified Knapp-Hartung keeps the variance inflation factor at
            # least one and uses a t reference distribution.  With only three
            # cohorts this is deliberately conservative and is reported beside
            # the conventional random-effects Wald interval.
            q_random = float(np.sum(wr * (g - est) ** 2))
            hk_scale = max(1.0, q_random / max(len(g) - 1, 1))
            hk_se = math.sqrt(hk_scale / float(np.sum(wr)))
            hk_df = len(g) - 1
            hk_crit = float(t.ppf(0.975, hk_df))
            return {
                "pooled_g": est,
                "se": se,
                "ci_low": est - 1.96 * se,
                "ci_high": est + 1.96 * se,
                "p": float(2 * norm.sf(abs(est / se))),
                "mKH_se": hk_se,
                "mKH_df": hk_df,
                "mKH_ci_low": est - hk_crit * hk_se,
                "mKH_ci_high": est + hk_crit * hk_se,
                "mKH_p": float(2 * t.sf(abs(est / hk_se), hk_df)),
                "tau2": tau2,
                "I2": i2,
                "Q": q,
            }

        result = pool(block)
        pooled_rows.append({"meta_group": label, "module": module, "k": len(block), **result})
        if len(block) >= 3:
            for omitted in block["dataset"]:
                sub = block[block["dataset"] != omitted]
                loo_rows.append({"meta_group": label, "module": module, "omitted_dataset": omitted, "k": len(sub), **pool(sub)})
    pooled = pd.DataFrame(pooled_rows)
    if not pooled.empty:
        pooled["FDR_BH"] = bh_fdr(pooled["p"])
        pooled["FDR_BH_mKH"] = bh_fdr(pooled["mKH_p"])
    return pooled, pd.DataFrame(loo_rows)


def cross_species_gene_concordance(
    mouse_payloads: list[dict[str, object]],
    human_payload: dict[str, object],
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Compare animal-level acute-EAE effects with human MS endothelial DGE.

    Mouse effects are calculated independently in each cohort and combined by
    random-effects meta-analysis.  Human effects come from the source study's
    EC-specific chronic-active-lesion versus control DESeq2 contrast.  Gene
    symbols are used as conservative one-to-one orthologue matches.  The local
    vascular-inflammatory spatial analysis has no deposited gene-level VI vs
    PPWM contrast, so it remains a separate module/pathway analysis.
    """
    effect_rows: list[dict[str, object]] = []
    for payload in mouse_payloads:
        genes = list(payload["genes"])
        expression = np.asarray(payload["expression"], dtype=float)
        groups = np.asarray(payload["groups"], dtype=object)
        case = str(payload["case"])
        control = str(payload["control"])
        case_idx = groups == case
        control_idx = groups == control
        # Duplicate symbols are uncommon; retain their average expression so a
        # gene contributes only once to the cross-species comparison.
        gene_to_rows: dict[str, list[int]] = {}
        for i, gene in enumerate(genes):
            gene_to_rows.setdefault(upper_gene(gene), []).append(i)
        for gene, idx in gene_to_rows.items():
            values = np.nanmean(expression[idx, :], axis=0)
            g, var_g, lo, hi = hedges_g(values[case_idx], values[control_idx])
            if np.isfinite(g) and np.isfinite(var_g):
                effect_rows.append(
                    {
                        "dataset": payload["dataset"],
                        "gene": gene,
                        "n_case": int(case_idx.sum()),
                        "n_control": int(control_idx.sum()),
                        "hedges_g": g,
                        "var_g": var_g,
                        "ci_low": lo,
                        "ci_high": hi,
                    }
                )
    mouse_effects = pd.DataFrame(effect_rows)
    common_mouse = set.intersection(
        *[set(mouse_effects.loc[mouse_effects["dataset"] == str(p["dataset"]), "gene"]) for p in mouse_payloads]
    )
    mouse_effects = mouse_effects[mouse_effects["gene"].isin(common_mouse)].copy()

    pooled_rows = []
    for gene, block in mouse_effects.groupby("gene"):
        g = block["hedges_g"].to_numpy(float)
        v = block["var_g"].to_numpy(float)
        w = 1.0 / v
        fixed = float(np.sum(w * g) / np.sum(w))
        q = float(np.sum(w * (g - fixed) ** 2))
        df = len(g) - 1
        c = float(np.sum(w) - np.sum(w * w) / np.sum(w))
        tau2 = max(0.0, (q - df) / c) if c > 0 else 0.0
        wr = 1.0 / (v + tau2)
        est = float(np.sum(wr * g) / np.sum(wr))
        se = math.sqrt(1.0 / float(np.sum(wr)))
        pooled_rows.append(
            {
                "gene": gene,
                "mouse_pooled_g": est,
                "mouse_se": se,
                "mouse_ci_low": est - 1.96 * se,
                "mouse_ci_high": est + 1.96 * se,
                "mouse_tau2": tau2,
                "mouse_I2": max(0.0, (q - df) / q) * 100.0 if q > 0 else 0.0,
                "mouse_same_direction_3of3": bool(np.all(np.sign(g) == np.sign(est))),
            }
        )
    pooled = pd.DataFrame(pooled_rows)

    human_genes = list(human_payload["genes"])
    human_expression = np.asarray(human_payload["expression"], dtype=float)
    human_groups = np.asarray(human_payload["groups"], dtype=object)
    human_case = human_groups == str(human_payload["case"])
    human_control = human_groups == str(human_payload["control"])
    human_rows = []
    human_gene_to_rows: dict[str, list[int]] = {}
    for i, gene in enumerate(human_genes):
        human_gene_to_rows.setdefault(upper_gene(gene), []).append(i)
    for gene, idx in human_gene_to_rows.items():
        values = np.nanmean(human_expression[idx, :], axis=0)
        g, var_g, lo, hi = hedges_g(values[human_case], values[human_control])
        if np.isfinite(g) and np.isfinite(var_g):
            human_rows.append(
                {
                    "gene": gene,
                    "human_hedges_g": g,
                    "human_var_g": var_g,
                    "human_ci_low": lo,
                    "human_ci_high": hi,
                    "human_n_case": int(human_case.sum()),
                    "human_n_control": int(human_control.sum()),
                }
            )
    human_effects = pd.DataFrame(human_rows)

    # Retain the source study's EC DESeq2 table as a validation column.  The
    # supplement reports a filtered set of DE genes, so it is not used as the
    # primary denominator for concordance.
    human_path = DATA / "GSE279183_Supplementary_Tables_1_10.xlsx"
    source_human = pd.read_excel(human_path, sheet_name="STable_6_Cell_type_DEG")
    source_human = source_human[
        (source_human["cell_type"] == "EC") & (source_human["contrast"] == "CAvsCtrl")
    ].copy()
    source_human["gene"] = source_human["Gene"].map(upper_gene)
    source_human = (
        source_human.groupby("gene", as_index=False)
        .agg(
            source_human_log2FC=("log2FoldChange", "mean"),
            source_human_Wald_stat=("stat", "mean"),
            source_human_p=("pvalue", "min"),
            source_human_FDR=("padj", "min"),
        )
    )
    merged = pooled.merge(human_effects, on="gene", how="inner").merge(source_human, on="gene", how="left")
    state_map: dict[str, list[str]] = {}
    for state, members in STATE_MODULES.items():
        for gene in members:
            state_map.setdefault(gene, []).append(state)
    merged["state"] = merged["gene"].map(lambda x: ";".join(state_map.get(x, [])))
    merged["prespecified_state_gene"] = merged["state"].ne("")
    merged["same_direction_mouse_human"] = (
        np.sign(merged["mouse_pooled_g"]) == np.sign(merged["human_hedges_g"])
    )
    merged["human_contrast"] = "GSE279183 reconstructed donor-level EC: chronic active MS vs control"
    merged["interpretation_boundary"] = (
        "gene-level source EC contrast; not the local VI-vs-PPWM spatial contrast"
    )

    def summarize(label: str, block: pd.DataFrame) -> dict[str, object]:
        block = block.replace([np.inf, -np.inf], np.nan).dropna(
            subset=["mouse_pooled_g", "human_hedges_g"]
        )
        rho, p_rho = spearmanr(block["mouse_pooled_g"], block["human_hedges_g"])
        n_same = int(block["same_direction_mouse_human"].sum())
        n = len(block)
        p_sign = float(binomtest(n_same, n, 0.5, alternative="two-sided").pvalue) if n else np.nan
        boot = []
        if 4 <= n <= 500:
            local_rng = np.random.default_rng(SEED + sum(map(ord, label)))
            x = block["mouse_pooled_g"].to_numpy(float)
            y = block["human_hedges_g"].to_numpy(float)
            for _ in range(10000):
                idx = local_rng.integers(0, n, n)
                if np.unique(x[idx]).size > 1 and np.unique(y[idx]).size > 1:
                    value = spearmanr(x[idx], y[idx]).statistic
                    if np.isfinite(value):
                        boot.append(float(value))
        if n > 500 and np.isfinite(rho) and abs(rho) < 1:
            z = np.arctanh(rho)
            se_z = 1 / math.sqrt(n - 3)
            rho_lo, rho_hi = np.tanh([z - 1.96 * se_z, z + 1.96 * se_z])
        else:
            rho_lo = float(np.quantile(boot, 0.025)) if boot else np.nan
            rho_hi = float(np.quantile(boot, 0.975)) if boot else np.nan
        return {
            "gene_set": label,
            "n_genes": n,
            "spearman_rho": float(rho) if np.isfinite(rho) else np.nan,
            "spearman_p": float(p_rho) if np.isfinite(p_rho) else np.nan,
            "rho_ci_low": rho_lo,
            "rho_ci_high": rho_hi,
            "same_direction_n": n_same,
            "direction_concordance": n_same / n if n else np.nan,
            "binomial_p_vs_0.5": p_sign,
            "median_mouse_g": float(block["mouse_pooled_g"].median()) if n else np.nan,
            "median_human_g": float(block["human_hedges_g"].median()) if n else np.nan,
        }

    primary = merged[merged["prespecified_state_gene"]].copy()
    summary_rows = [summarize("All_prespecified_state_genes", primary)]
    for state in STATE_MODULES:
        summary_rows.append(summarize(state, merged[merged["state"].str.contains(state, regex=False)]))
    summary_rows.append(summarize("All_common_genes_sensitivity", merged))
    concordance = pd.DataFrame(summary_rows)
    concordance["FDR_BH_spearman_prespecified_tests"] = np.nan
    primary_test = concordance["gene_set"].ne("All_common_genes_sensitivity")
    concordance.loc[primary_test, "FDR_BH_spearman_prespecified_tests"] = bh_fdr(
        concordance.loc[primary_test, "spearman_p"]
    )
    source_valid = merged.dropna(subset=["source_human_Wald_stat", "source_human_log2FC"]).copy()
    mouse_source = spearmanr(source_valid["mouse_pooled_g"], source_valid["source_human_Wald_stat"])
    reconstructed_source = spearmanr(source_valid["human_hedges_g"], source_valid["source_human_Wald_stat"])
    source_sensitivity = pd.DataFrame(
        [
            {
                "n_reported_source_EC_DEGs": len(source_valid),
                "mouse_meta_vs_source_Wald_rho": float(mouse_source.statistic),
                "mouse_meta_vs_source_Wald_p": float(mouse_source.pvalue),
                "mouse_vs_source_direction_concordance": float(
                    (np.sign(source_valid["mouse_pooled_g"]) == np.sign(source_valid["source_human_log2FC"])).mean()
                ),
                "reconstructed_human_vs_source_Wald_rho": float(reconstructed_source.statistic),
                "reconstructed_human_vs_source_Wald_p": float(reconstructed_source.pvalue),
                "reconstructed_human_vs_source_direction_concordance": float(
                    (np.sign(source_valid["human_hedges_g"]) == np.sign(source_valid["source_human_log2FC"])).mean()
                ),
                "boundary": "source supplement contains a filtered EC DEG set and is used only as sensitivity validation",
            }
        ]
    )
    return (
        mouse_effects,
        merged.sort_values(["prespecified_state_gene", "state", "gene"], ascending=[False, True, True]),
        concordance,
        source_sensitivity,
    )


def analyze_vegf(scores: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for module in ALL_MODULES:
        cfa = scores.loc[scores["group"] == "CFA", module].dropna().to_numpy(float)
        acute = scores.loc[scores["group"] == "Acute", module].dropna().to_numpy(float)
        treated = scores.loc[scores["group"] == "Bevacizumab", module].dropna().to_numpy(float)
        if min(len(cfa), len(acute), len(treated)) < 2:
            continue
        disease = float(np.mean(acute) - np.mean(cfa))
        treatment, p = exact_mean_permutation(treated, acute)
        residual = float(np.mean(treated) - np.mean(cfa))
        if abs(treatment) < 0.25 * max(abs(disease), 1e-12):
            classification = "little_change"
        elif np.sign(treatment) == -np.sign(disease) and abs(residual) < abs(disease):
            classification = "near_normalization" if abs(residual) <= 0.25 * abs(disease) else "partial_attenuation"
        elif np.sign(treatment) == np.sign(disease):
            classification = "discordant"
        else:
            classification = "uncertain"
        rows.append(
            {
                "module": module,
                "n_CFA": len(cfa),
                "n_acute": len(acute),
                "n_bevacizumab": len(treated),
                "acute_minus_CFA": disease,
                "bevacizumab_minus_acute": treatment,
                "bevacizumab_minus_CFA": residual,
                "p_exact_treatment": p,
                "classification": classification,
            }
        )
    out = pd.DataFrame(rows)
    out["FDR_BH_treatment"] = bh_fdr(out["p_exact_treatment"])
    return out


def main() -> None:
    ensure_out()
    sample_scores = []
    coverage = []
    qc = []

    g210, cov210, qc210, gene210 = analyze_gse210776()
    g199, cov199, qc199, gene199 = analyze_gse199460()
    g954, cov954, qc954, gene954 = analyze_gse95401()
    g279, cov279, qc279, gene279 = analyze_gse279183()
    g208, cov208, qc208 = analyze_gse208747()

    sample_scores.extend([g210, g199, g954, g279, g208])
    coverage.extend([cov210, cov199, cov954, cov279, cov208])
    qc.extend([qc210, qc199, qc954, qc279, qc208])

    score_table = pd.concat(sample_scores, ignore_index=True, sort=False)
    coverage_table = pd.concat(coverage, ignore_index=True, sort=False)
    qc_table = pd.concat(qc, ignore_index=True, sort=False)
    score_table.to_csv(OUT / "sample_level_module_scores.csv", index=False)
    coverage_table.to_csv(OUT / "module_gene_coverage.csv", index=False)
    qc_table.to_csv(OUT / "reconstruction_qc.csv", index=False)

    effects = []
    effects += contrast_rows(g210, "group", "Acute", "CFA", "GSE210776", "Acute_EAE_vs_CFA")
    effects += contrast_rows(g210, "group", "Chronic", "CFA", "GSE210776", "Chronic_EAE_vs_CFA")
    g199_all = g199[g199["subset"] == "All_endothelial"]
    g199_venous = g199[g199["subset"] == "Venous_like"]
    effects += contrast_rows(g199_all, "group", "Acute", "Control", "GSE199460_all", "Acute_EAE_vs_control")
    effects += contrast_rows(g199_venous, "group", "Acute", "Control", "GSE199460_venous", "Acute_EAE_vs_control")
    effects += contrast_rows(g954, "group", "Acute", "Control", "GSE95401", "Acute_EAE_vs_control")
    effects += contrast_rows(g954, "group", "Subacute", "Control", "GSE95401", "Subacute_EAE_vs_control")
    effects += contrast_rows(g954, "group", "Chronic", "Control", "GSE95401", "Chronic_EAE_vs_control")
    effects += contrast_rows(g279, "group", "Chronic_active", "Control", "GSE279183", "Chronic_active_MS_vs_control")
    effects += contrast_rows(g279, "group", "Chronic_inactive", "Control", "GSE279183", "Chronic_inactive_MS_vs_control")
    effects += contrast_rows(
        g279,
        "group",
        "Chronic_active",
        "Chronic_inactive",
        "GSE279183",
        "Chronic_active_vs_chronic_inactive",
    )
    effects += contrast_rows(g208, "group", "Active", "NAWM", "GSE208747", "Active_lesion_vs_NAWM")
    effects += contrast_rows(g208, "group", "Mixed_active_inactive", "NAWM", "GSE208747", "Mixed_lesion_vs_NAWM")
    effects += contrast_rows(g208, "group", "Active", "Mixed_active_inactive", "GSE208747", "Active_vs_mixed_lesion")
    effect_table = pd.DataFrame(effects)
    effect_table.to_csv(OUT / "dataset_module_effects.csv", index=False)

    # Keep the lesion-stage analysis anchored to the deposited, donor-level,
    # endothelial-abundance-adjusted sensitivity table.  These contrasts are
    # descriptive because the stage groups are small and no q30 comparison
    # survives FDR correction.
    stage_source = pd.read_csv(SOURCE / "S10_GSE208747.csv")
    stage_source = stage_source[
        (stage_source["threshold"] == "q30")
        & (stage_source["value"] == "module_z_ECadj")
        & stage_source["contrast"].isin(["Active_vs_NAWM", "Mixed_vs_NAWM", "Active_vs_Mixed"])
    ].copy()
    stage_source.to_csv(OUT / "GSE208747_stage_locked_modules_q30_ECadj.csv", index=False)

    acute_meta_input = effect_table[
        effect_table["contrast"].isin(["Acute_EAE_vs_CFA", "Acute_EAE_vs_control"])
        & effect_table["dataset"].isin(["GSE210776", "GSE199460_all", "GSE95401"])
    ]
    acute_meta, acute_loo = random_effects_meta(acute_meta_input, "Acute_EAE")
    acute_meta.to_csv(OUT / "acute_EAE_random_effects_meta.csv", index=False)
    acute_loo.to_csv(OUT / "acute_EAE_leave_one_dataset_out.csv", index=False)

    mouse_gene_effects, cross_species, cross_summary, cross_source_sensitivity = cross_species_gene_concordance(
        [gene210, gene199, gene954], gene279
    )
    mouse_gene_effects.to_csv(OUT / "acute_EAE_gene_effects_by_dataset.csv", index=False)
    cross_species.to_csv(OUT / "cross_species_gene_effects.csv", index=False)
    cross_summary.to_csv(OUT / "cross_species_concordance_summary.csv", index=False)
    cross_source_sensitivity.to_csv(OUT / "cross_species_source_DGE_sensitivity.csv", index=False)

    stage_effects = effect_table[
        effect_table["contrast"].isin(
            [
                "Acute_EAE_vs_CFA", "Chronic_EAE_vs_CFA", "Subacute_EAE_vs_control",
                "Chronic_EAE_vs_control", "Active_lesion_vs_NAWM", "Mixed_lesion_vs_NAWM",
                "Chronic_active_MS_vs_control", "Chronic_inactive_MS_vs_control",
                "Chronic_active_vs_chronic_inactive",
            ]
        )
    ].copy()
    stage_effects.to_csv(OUT / "stage_resolved_effects.csv", index=False)

    vegf = analyze_vegf(g210)
    vegf.to_csv(OUT / "VEGFA_three_arm_transcriptional_perturbation.csv", index=False)


    # Infection datasets are retained only as an auditable supplementary table.
    # They do not measure brain endothelium and are excluded from vascular results.
    infection_context = pd.read_csv(SOURCE / "S13_infection_context.csv")
    infection_context.to_csv(OUT / "Table_S13_infection_context.csv", index=False)

    frozen_vegf = pd.read_csv(SOURCE / "S7_BEVAC3.csv")
    frozen_vegf.to_csv(OUT / "VEGFA_venous_source_effects.csv", index=False)
    spatial_niche = pd.read_csv(SOURCE / "S9_GSE279183_spatial.csv")
    spatial_niche.to_csv(OUT / "GSE279183_inflammatory_vascular_microenvironment_source_effects.csv", index=False)
    spatial_barrier = pd.DataFrame(
        [
            {"module": "Tight_junction", "delta_microenvironment_vs_control_WM": -0.588, "p": 0.0177, "FDR_BH": 0.0435},
            {"module": "BBB_transport_identity", "delta_microenvironment_vs_control_WM": -0.396, "p": 0.2096, "FDR_BH": 0.2795},
            {"module": "Wnt_BBB", "delta_microenvironment_vs_control_WM": -1.375, "p": 0.0013, "FDR_BH": 0.0045},
        ]
    )
    spatial_barrier["source"] = "S14 verified audit of Integrated_human_spatial_barrier_identity.csv"
    spatial_barrier.to_csv(OUT / "GSE279183_inflammatory_vascular_microenvironment_barrier_effects.csv", index=False)

    summary = {
        "seed": SEED,
        "primary_unit": "animal or donor",
        "acute_EAE_datasets_in_meta": sorted(acute_meta_input["dataset"].unique().tolist()),
        "gse210776_raw_droplet_filter": "UMI>=300; genes>=150; >=2 EC markers; EC score greater than mural/glial/immune score",
        "gse279183_endothelial_nuclei": int(qc279["endothelial_nuclei"].sum()),
        "gse208747_sections": int(qc208["sample_id"].nunique()),
        "cross_species_primary_comparison": "mouse acute-EAE pooled gene effects vs reconstructed donor-level human EC chronic-active-MS-vs-control effects",
        "cross_species_spatial_boundary": "no deposited gene-level inflammatory-vascular-microenvironment-vs-PPWM contrast; spatial evidence remains module/pathway level",
        "meta_small_study_sensitivity": "modified Knapp-Hartung intervals reported beside conventional random-effects Wald intervals",
        "infection_datasets": "supplementary table only; excluded from vascular conclusions",
        "interpretation": "RNA changes do not establish functional BBB recovery",
    }
    with open(OUT / "analysis_summary.json", "w", encoding="utf-8") as fh:
        json.dump(summary, fh, ensure_ascii=False, indent=2)
    print(json.dumps({"status": "completed", "output": str(OUT), **summary}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
