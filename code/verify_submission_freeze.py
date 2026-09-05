#!/usr/bin/env python3
"""Verify package completeness and agreement with the frozen manuscript."""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import pandas as pd
from openpyxl import load_workbook
from scipy.stats import spearmanr


ROOT = Path(__file__).resolve().parents[1]
TITLE = "Brain endothelial responses in acute EAE partially overlap with vascular changes in multiple sclerosis"
SEED = 20260901


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def docx_text_and_media_hashes(path: Path) -> tuple[str, set[str]]:
    with zipfile.ZipFile(path) as archive:
        root = ET.fromstring(archive.read("word/document.xml"))
        text = "\n".join(node.text or "" for node in root.iter() if node.tag.endswith("}t"))
        hashes = {
            hashlib.sha256(archive.read(name)).hexdigest()
            for name in archive.namelist()
            if name.startswith("word/media/")
        }
    return text, hashes


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--manuscript", type=Path)
    parser.add_argument("--report", type=Path)
    parser.add_argument("--check-manifest", action="store_true",
                        help="Check the unmodified downloaded archive before regenerating figures.")
    args = parser.parse_args()

    checks: list[dict[str, object]] = []

    def check(name: str, passed: bool, detail: str) -> None:
        checks.append({"check": name, "passed": bool(passed), "detail": detail})

    required = [
        "README.md",
        "public_data/README.md",
        "REPRODUCIBILITY.md",
        "code/run_strict_orthology.py",
        "MANIFEST.sha256",
        ".zenodo.json",
        "CITATION.cff",
        "environment.yml",
        "requirements.txt",
        "software_versions.csv",
        "analysis_parameters.json",
        "results/analysis_summary.json",
        "supplementary_tables/Supplementary_Tables_S1-S18.xlsx",
        "figures/BBI_Supplementary_Figures_S1-S9.pdf",
        "figure_source_data/figure_source_index.csv",
    ]
    missing = [item for item in required if not (ROOT / item).is_file()]
    check("required files", not missing, "missing: " + ", ".join(missing) if missing else "all required files present")

    legacy = [name for name in ("analysis", "analysis_results") if (ROOT / name).exists()]
    check("legacy directories removed", not legacy, ", ".join(legacy) if legacy else "none")
    metadata = json.loads((ROOT / ".zenodo.json").read_text(encoding="utf-8"))
    citation = (ROOT / "CITATION.cff").read_text(encoding="utf-8")
    doi_match = re.search(r'^doi:\s*[\"\x27]?([^\"\x27\s]+)', citation, re.MULTILINE)
    doi = "10.5281/zenodo.22340814"
    check("version DOI metadata", metadata.get("doi") == doi and doi_match is not None and doi_match.group(1) == doi,
          doi + " (reserved during package preparation)")

    if args.check_manifest:
        entries = {}
        malformed = []
        for line in (ROOT / "MANIFEST.sha256").read_text(encoding="utf-8").splitlines():
            match = re.fullmatch(r"([0-9a-f]{64})  (.+)", line)
            if not match:
                malformed.append(line)
                continue
            digest, name = match.groups()
            relative = Path(name)
            if relative.is_absolute() or ".." in relative.parts or name in entries:
                malformed.append(line)
                continue
            entries[name] = digest
        bad = [name for name, digest in entries.items()
               if not (ROOT / name).is_file() or sha256(ROOT / name) != digest]
        actual = {p.relative_to(ROOT).as_posix() for p in ROOT.rglob("*")
                  if p.is_file() and not any(part in {".git", "__pycache__"} for part in p.relative_to(ROOT).parts)
                  and p.name != ".DS_Store" and p.suffix != ".pyc"
                  and (args.report is None or p.resolve() != args.report.resolve())}
        extras = sorted(actual - set(entries) - {"MANIFEST.sha256"})
        check("archive manifest", not malformed and not bad and not extras,
              f"{len(entries)} entries verified" if not malformed and not bad and not extras
              else f"malformed={malformed}, missing/changed={bad}, extra={extras}")

    main_figures = sorted((ROOT / "figures" / "main").glob("Figure_*.png"))
    supplementary_figures = sorted((ROOT / "figures" / "supplementary").glob("Figure_S*.png"))
    check("main figure count", len(main_figures) == 5, f"found {len(main_figures)}")
    check("supplementary figure count", len(supplementary_figures) == 9, f"found {len(supplementary_figures)}")

    table_files = list((ROOT / "supplementary_tables").glob("S*.csv"))
    represented = set()
    for path in table_files:
        token = path.name.split("_", 1)[0]
        if token.startswith("S") and token[1:].isdigit():
            represented.add(int(token[1:]))
    check("supplementary tables S1 to S18", represented == set(range(1, 19)), f"represented tables: {sorted(represented)}")

    workbook_path = ROOT / "supplementary_tables" / "Supplementary_Tables_S1-S18.xlsx"
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    expected_sheets = {
        "README", "S1_Datasets", "S2_Signatures", "S3_EAE_modules", "S4_EAE_genes",
        "S5_Mouse_sender", "S6_EAE_time", "S7_VEGFA_venous", "S8_MS_snRNA",
        "S9_MS_spatial", "S10_Lesion_stage", "S11_Viral_context", "S12_Bacterial_context",
        "S13_Infection_context", "S14_Result_audit", "S15_Orthology", "S16_MERFISH",
        "S17_Eligibility", "S18_Gene_sets", "S18_Coverage",
    }
    check("combined workbook sheets", set(workbook.sheetnames) == expected_sheets, f"found {len(workbook.sheetnames)} sheets")
    formula_errors = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}:
                    formula_errors.append(f"{sheet.title}!{cell.coordinate}={cell.value}")
    check("workbook formula errors", not formula_errors, ", ".join(formula_errors) if formula_errors else "none")

    parameters = json.loads((ROOT / "analysis_parameters.json").read_text(encoding="utf-8"))
    check("random seed", parameters.get("random_seed") == SEED, str(parameters.get("random_seed")))
    check("orthology bootstrap seed", parameters.get("orthology_bootstrap_seed") == 20260904
          and parameters.get("orthology_bootstrap_resamples") == 2000,
          "20260904; 2000 resamples, as used by the original S15 analysis")
    figure_script = (ROOT / "code" / "make_bbi_figures.py").read_text(encoding="utf-8")
    check("selected Figure 4 background", "background_color: str = '#BFC4C8'" in figure_script and "background_alpha: float = .48" in figure_script, "light gray settings present")

    stale_names = [
        str(path.relative_to(ROOT))
        for path in ROOT.rglob("*")
        if path.is_file() and any(token in path.name.lower() for token in ["candidate", "figure_s10", "ms_gwas"])
    ]
    check("stale artifact filenames", not stale_names, ", ".join(stale_names) if stale_names else "none")

    stale_seed_files = []
    text_suffixes = {".py", ".md", ".csv", ".json", ".yml", ".txt", ".cff"}
    for path in ROOT.rglob("*"):
        if path.is_file() and path.suffix.lower() in text_suffixes:
            try:
                obsolete_seed = "2026" + "0820"
                if obsolete_seed in path.read_text(encoding="utf-8"):
                    stale_seed_files.append(str(path.relative_to(ROOT)))
            except UnicodeDecodeError:
                pass
    check("obsolete random seed", not stale_seed_files, ", ".join(stale_seed_files) if stale_seed_files else "none")

    meta = pd.read_csv(ROOT / "results" / "acute_EAE_random_effects_meta.csv").set_index("module")
    check("acute EAE pooled immune effect", math.isclose(meta.loc["Endothelial_immune_activation", "pooled_g"], 2.596911, abs_tol=1e-6), f"{meta.loc['Endothelial_immune_activation', 'pooled_g']:.6f}")
    check("acute EAE pooled BBB effect", math.isclose(meta.loc["BBB_specialization", "pooled_g"], -1.790610, abs_tol=1e-6), f"{meta.loc['BBB_specialization', 'pooled_g']:.6f}")
    check("modified Knapp Hartung intervals cross zero", meta.loc["Endothelial_immune_activation", "mKH_ci_low"] < 0 < meta.loc["Endothelial_immune_activation", "mKH_ci_high"] and meta.loc["BBB_specialization", "mKH_ci_low"] < 0 < meta.loc["BBB_specialization", "mKH_ci_high"], "both principal intervals cross zero")

    cross = pd.read_csv(ROOT / "results" / "cross_species_concordance_summary.csv").set_index("gene_set")
    row = cross.loc["All_common_genes_sensitivity"]
    check("cross species all gene result", int(row.n_genes) == 11633 and math.isclose(row.spearman_rho, 0.1972107329, abs_tol=1e-9), f"n={int(row.n_genes)}, rho={row.spearman_rho:.6f}")
    row = cross.loc["All_prespecified_state_genes"]
    check("cross species focused result", int(row.n_genes) == 43 and math.isclose(row.spearman_rho, 0.243431, abs_tol=1e-6), f"n={int(row.n_genes)}, rho={row.spearman_rho:.6f}")

    strict = pd.read_csv(ROOT / "results" / "strict_orthology_sensitivity_summary.csv")
    row = strict[(strict.mapping.str.startswith("HCOP")) & (strict.scope == "All shared genes")].iloc[0]
    check("strict orthology all gene result", int(row.n_genes) == 7705 and math.isclose(row.spearman_rho, 0.1868056122, abs_tol=1e-9), f"n={int(row.n_genes)}, rho={row.spearman_rho:.6f}")

    # Validate the archived correlations against gene effects, independently of
    # the summary tables. This does not repeat filtering of the original HCOP export.
    gene_tables = {
        "Uppercase-symbol match": pd.read_csv(ROOT / "results" / "cross_species_gene_effects.csv"),
        "HCOP reciprocal 1:1; Ensembl+NCBI support": pd.read_csv(ROOT / "results" / "strict_orthology_gene_effects.csv"),
    }
    concordance_errors = []
    for item in strict.itertuples():
        genes = gene_tables[item.mapping]
        if item.scope == "Prespecified focused genes":
            genes = genes[genes.prespecified_state_gene.astype(str).str.lower().eq("true")]
        rho = float(spearmanr(genes.mouse_pooled_g, genes.human_hedges_g).statistic)
        direction = float(((genes.mouse_pooled_g > 0) & (genes.human_hedges_g > 0)
                          | (genes.mouse_pooled_g < 0) & (genes.human_hedges_g < 0)).mean())
        if (len(genes) != item.n_genes or not math.isclose(rho, item.spearman_rho, abs_tol=1e-12)
                or not math.isclose(direction, item.same_direction_fraction, abs_tol=1e-12)):
            concordance_errors.append(f"{item.mapping}: {item.scope}")
    check("concordance recomputed from frozen gene effects", not concordance_errors,
          ", ".join(concordance_errors) if concordance_errors else "all four counts, correlations and direction fractions match")

    merfish = pd.read_csv(ROOT / "results" / "GSE284005_paired_summary.csv")
    row = merfish[merfish.feature == "Stress/inflammatory endothelial fraction"].iloc[0]
    check("MERFISH paired composition", int(row.donors_higher_in_Vas_Imm) == 5 and math.isclose(row.exact_two_sided_sign_flip_p, 0.0625, abs_tol=1e-12), f"{int(row.donors_higher_in_Vas_Imm)} of 5, P={row.exact_two_sided_sign_flip_p}")

    functional = pd.read_csv(ROOT / "results" / "published_source_barrier_and_vascular_summary.csv").set_index("endpoint")
    check("IgG leakage result", math.isclose(functional.loc["IgG leakage", "p_exact_two_sided"], 0.373385, abs_tol=1e-6), f"P={functional.loc['IgG leakage', 'p_exact_two_sided']:.6f}")
    check("fibrinogen leakage result", math.isclose(functional.loc["Fibrinogen leakage", "p_exact_two_sided"], 0.913945, abs_tol=1e-6), f"P={functional.loc['Fibrinogen leakage', 'p_exact_two_sided']:.6f}")

    source_index = pd.read_csv(ROOT / "figure_source_data" / "figure_source_index.csv")
    bad_source_rows = []
    for row in source_index.itertuples():
        source = ROOT / row.source_file
        if not source.is_file() or sha256(source) != row.sha256:
            bad_source_rows.append(f"{row.figure}: {row.source_file}")
    check("figure source index", not bad_source_rows, ", ".join(bad_source_rows) if bad_source_rows else f"{len(source_index)} mappings verified")

    if args.manuscript:
        manuscript_text, media_hashes = docx_text_and_media_hashes(args.manuscript)
        check("manuscript title", TITLE in manuscript_text, TITLE)
        missing_embedded = [path.name for path in main_figures if sha256(path) not in media_hashes]
        check("main figures embedded in manuscript", not missing_embedded, ", ".join(missing_embedded) if missing_embedded else "all five PNG hashes found in the DOCX media")

    passed = all(item["passed"] for item in checks)
    report = {"package": str(ROOT), "passed": passed, "checks": checks}
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(json.dumps(report, indent=2))
    raise SystemExit(0 if passed else 1)


if __name__ == "__main__":
    main()
